import io
import logging
import os
from datetime import datetime, timezone
from typing import List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from dependencies.auth import get_current_user
from models.debts import Debts
from models.payments import Payments
from models.payment_details import PaymentDetails
from schemas.auth import UserResponse
from services.debts import DebtsService
from services.permission_check import require_permission, resolve_user_role

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/debts", tags=["debts"])


# ---------- Pydantic Schemas ----------
class DebtData(BaseModel):
    customer_name: str
    invoice_number: str
    amount: float
    invoice_date: Optional[str] = None
    notes: Optional[str] = None


class DebtResponse(BaseModel):
    id: int
    user_id: str
    customer_name: str
    invoice_number: str
    amount: float
    remaining_amount: float
    invoice_date: Optional[str] = None
    status: str
    notes: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class DebtListResponse(BaseModel):
    items: List[DebtResponse]
    total: int
    skip: int
    limit: int


class PaymentDebtItem(BaseModel):
    debt_id: int
    amount_applied: float


class PaymentRequest(BaseModel):
    customer_name: str
    debts: List[PaymentDebtItem] = []
    manual_amount: Optional[float] = None  # used when debts is empty: a payment not tied to a specific invoice
    discount: float = 0
    returns_amount: float = 0
    notes: Optional[str] = None


class PaymentResponse(BaseModel):
    id: int
    user_id: str
    customer_name: str
    total_selected: float
    discount: float
    returns_amount: float
    net_received: float
    notes: Optional[str] = None
    created_at: Optional[datetime] = None
    rep_name: Optional[str] = None
    receipt_captured_at: Optional[datetime] = None
    status: Optional[str] = None
    canceled_at: Optional[datetime] = None
    canceled_by: Optional[str] = None
    cancel_reason: Optional[str] = None
    handed_over: Optional[bool] = None
    handed_over_at: Optional[datetime] = None
    handed_over_by: Optional[str] = None

    class Config:
        from_attributes = True


class ConfirmHandoverRequest(BaseModel):
    payment_ids: List[int]


class HandoverRepSummary(BaseModel):
    user_id: str
    rep_name: str
    pending_total: float
    payments: List[PaymentResponse]


class HandoverSummaryResponse(BaseModel):
    reps: List[HandoverRepSummary]
    pending_total: float
    confirmed_total: float


class ReceiptImageUpload(BaseModel):
    image_base64: str


MAX_RECEIPT_IMAGE_CHARS = 6_000_000  # ~4.3MB decoded; generous cap for a compressed JPEG


class CancelPaymentRequest(BaseModel):
    reason: Optional[str] = None


class PaymentListResponse(BaseModel):
    items: List[PaymentResponse]
    total: int
    skip: int
    limit: int


class ImportResult(BaseModel):
    total_rows: int
    imported: int
    skipped: int
    errors: List[str]


# ---------- Customer Debt Summary ----------
class CustomerDebtSummary(BaseModel):
    customer_name: str
    total_debts: int
    unpaid_count: int
    partial_count: int
    paid_count: int
    total_amount: float
    total_remaining: float
    status: str  # clean, has_debt, overdue


@router.get("/customer-summary/{customer_name}", response_model=CustomerDebtSummary)
async def get_customer_debt_summary(
    customer_name: str,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Get debt summary for a specific customer name"""
    from sqlalchemy import select, func

    # Get counts by status
    result = await db.execute(
        select(
            func.count(Debts.id).label("total"),
            func.coalesce(func.sum(Debts.amount), 0).label("total_amount"),
            func.coalesce(func.sum(Debts.remaining_amount), 0).label("total_remaining"),
            func.count(Debts.id).filter(Debts.status == "unpaid").label("unpaid_count"),
            func.count(Debts.id).filter(Debts.status == "partial").label("partial_count"),
            func.count(Debts.id).filter(Debts.status == "paid").label("paid_count"),
        ).where(Debts.customer_name == customer_name)
    )
    row = result.one()

    total_remaining = float(row.total_remaining or 0)
    unpaid_count = int(row.unpaid_count or 0)
    partial_count = int(row.partial_count or 0)

    if total_remaining <= 0 and int(row.total or 0) == 0:
        status = "clean"
    elif unpaid_count > 0 or partial_count > 0:
        status = "has_debt"
    else:
        status = "clean"

    return CustomerDebtSummary(
        customer_name=customer_name,
        total_debts=int(row.total or 0),
        unpaid_count=unpaid_count,
        partial_count=partial_count,
        paid_count=int(row.paid_count or 0),
        total_amount=float(row.total_amount or 0),
        total_remaining=total_remaining,
        status=status,
    )


@router.get("/customer-summaries", response_model=List[CustomerDebtSummary])
async def get_all_customer_summaries(
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Get debt summaries for all customers that have any debts"""
    from sqlalchemy import select, func

    result = await db.execute(
        select(
            Debts.customer_name,
            func.count(Debts.id).label("total"),
            func.coalesce(func.sum(Debts.amount), 0).label("total_amount"),
            func.coalesce(func.sum(Debts.remaining_amount), 0).label("total_remaining"),
            func.count(Debts.id).filter(Debts.status == "unpaid").label("unpaid_count"),
            func.count(Debts.id).filter(Debts.status == "partial").label("partial_count"),
            func.count(Debts.id).filter(Debts.status == "paid").label("paid_count"),
        )
        .group_by(Debts.customer_name)
        .order_by(Debts.customer_name)
    )
    rows = result.all()

    summaries = []
    for row in rows:
        total_remaining = float(row.total_remaining or 0)
        unpaid_count = int(row.unpaid_count or 0)
        partial_count = int(row.partial_count or 0)

        if total_remaining <= 0:
            status = "clean"
        elif unpaid_count > 0 or partial_count > 0:
            status = "has_debt"
        else:
            status = "clean"

        summaries.append(CustomerDebtSummary(
            customer_name=row.customer_name,
            total_debts=int(row.total or 0),
            unpaid_count=unpaid_count,
            partial_count=partial_count,
            paid_count=int(row.paid_count or 0),
            total_amount=float(row.total_amount or 0),
            total_remaining=total_remaining,
            status=status,
        ))

    return summaries


# ---------- Debt CRUD Routes ----------
@router.get("", response_model=DebtListResponse)
async def list_debts(
    query: str = Query(None),
    sort: str = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=2000),
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """List all debts with filtering and pagination"""
    import json as json_mod

    service = DebtsService(db)
    query_dict = None
    if query:
        try:
            query_dict = json_mod.loads(query)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid query JSON")

    result = await service.get_list(skip=skip, limit=limit, query_dict=query_dict, sort=sort)
    return result


@router.get("/customers", response_model=List[str])
async def get_customer_names(
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Get distinct customer names with debts"""
    service = DebtsService(db)
    return await service.get_customer_names()


@router.get("/unpaid/{customer_name}", response_model=List[DebtResponse])
async def get_unpaid_debts(
    customer_name: str,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Get all unpaid/partial debts for a customer"""
    service = DebtsService(db)
    return await service.get_unpaid_by_customer(customer_name)


# ---------- Template Download (MUST be before /{debt_id} to avoid route conflict) ----------
@router.get("/template")
async def download_template():
    """Download the Excel template for debt import"""
    # Generate template dynamically using openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "قالب الديون"
        ws.sheet_view.rightToLeft = True

        # Headers
        headers = ["اسم الزبون", "رقم القائمة", "مبلغ الدين", "تاريخ القائمة"]
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Sample data rows
        sample_data = [
            ["صيدلية النور", "INV-001", 500000, "2025-01-15"],
            ["صيدلية الأمل", "INV-002", 750000, "2025-02-01"],
            ["صيدلية الشفاء", "INV-003", 320000, "2025-03-10"],
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=debt_template.xlsx",
                "Cache-Control": "no-cache",
            },
        )
    except ImportError:
        # Fallback: generate with pandas
        df = pd.DataFrame({
            "اسم الزبون": ["صيدلية النور", "صيدلية الأمل", "صيدلية الشفاء"],
            "رقم القائمة": ["INV-001", "INV-002", "INV-003"],
            "مبلغ الدين": [500000, 750000, 320000],
            "تاريخ القائمة": ["2025-01-15", "2025-02-01", "2025-03-10"],
        })
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, sheet_name="قالب الديون")
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=debt_template.xlsx",
                "Cache-Control": "no-cache",
            },
        )


@router.post("", response_model=DebtResponse, status_code=201)
async def create_debt(
    data: DebtData,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Create a single debt record"""
    service = DebtsService(db)
    existing = await service.get_by_invoice_number(data.invoice_number)
    if existing:
        raise HTTPException(status_code=400, detail=f"رقم القائمة {data.invoice_number} موجود مسبقاً")

    result = await service.create(data.model_dump(), user_id=str(current_user.id))
    return result


@router.delete("/{debt_id}")
async def delete_debt(
    debt_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Delete a debt record (requires can_delete permission on the debts page)"""
    await require_permission(db, current_user, "debts", "delete")
    service = DebtsService(db)
    success = await service.delete(debt_id)
    if not success:
        raise HTTPException(status_code=404, detail="الدين غير موجود")
    return {"message": "تم حذف الدين بنجاح"}


# ---------- Excel Import ----------
@router.post("/import-excel", response_model=ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Import debts from Excel file.
    Expected columns: اسم الزبون, رقم القائمة, مبلغ الدين, تاريخ القائمة
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="يرجى رفع ملف Excel (.xlsx أو .xls)")

    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"خطأ في قراءة الملف: {str(e)}")

    # Normalize column names (strip whitespace)
    df.columns = [str(c).strip() for c in df.columns]

    # Map expected Arabic column names
    col_map = {
        "اسم الزبون": "customer_name",
        "رقم القائمة": "invoice_number",
        "مبلغ الدين": "amount",
        "تاريخ القائمة": "invoice_date",
    }

    # Check required columns
    missing = [k for k in ["اسم الزبون", "رقم القائمة", "مبلغ الدين"] if k not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"أعمدة مفقودة في الملف: {', '.join(missing)}. الأعمدة المطلوبة: اسم الزبون، رقم القائمة، مبلغ الدين، تاريخ القائمة",
        )

    df = df.rename(columns=col_map)

    service = DebtsService(db)
    imported = 0
    skipped = 0
    errors: List[str] = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row (1-indexed + header)
        try:
            customer_name = str(row.get("customer_name", "")).strip()
            invoice_number = str(row.get("invoice_number", "")).strip()
            amount_raw = row.get("amount", 0)

            if not customer_name or not invoice_number:
                errors.append(f"صف {row_num}: اسم الزبون أو رقم القائمة فارغ")
                skipped += 1
                continue

            try:
                amount = float(amount_raw)
            except (ValueError, TypeError):
                errors.append(f"صف {row_num}: مبلغ الدين غير صالح ({amount_raw})")
                skipped += 1
                continue

            if amount <= 0:
                errors.append(f"صف {row_num}: مبلغ الدين يجب أن يكون أكبر من صفر")
                skipped += 1
                continue

            # Check for duplicate invoice number
            existing = await service.get_by_invoice_number(invoice_number)
            if existing:
                errors.append(f"صف {row_num}: رقم القائمة {invoice_number} موجود مسبقاً - تم تخطيه")
                skipped += 1
                continue

            # Parse date
            invoice_date = None
            date_raw = row.get("invoice_date")
            if pd.notna(date_raw):
                if isinstance(date_raw, datetime):
                    invoice_date = date_raw.strftime("%Y-%m-%d")
                else:
                    invoice_date = str(date_raw).strip()

            await service.create(
                {
                    "customer_name": customer_name,
                    "invoice_number": invoice_number,
                    "amount": amount,
                    "remaining_amount": amount,
                    "invoice_date": invoice_date,
                    "status": "unpaid",
                },
                user_id=str(current_user.id),
            )
            imported += 1

        except Exception as e:
            errors.append(f"صف {row_num}: خطأ - {str(e)}")
            skipped += 1

    return ImportResult(
        total_rows=len(df),
        imported=imported,
        skipped=skipped,
        errors=errors[:20],  # Limit error messages
    )


# ---------- Payment Routes ----------
@router.post("/pay", response_model=PaymentResponse)
async def process_payment(
    data: PaymentRequest,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Process a payment for selected debts.
    Calculates: net_received = total_selected - discount - returns_amount
    Updates debt statuses accordingly.
    """
    service = DebtsService(db)

    debt_objects = []
    if data.debts:
        # Validate all debts exist and belong to the customer
        total_selected = 0.0
        for item in data.debts:
            debt = await service.get_by_id(item.debt_id)
            if not debt:
                raise HTTPException(status_code=404, detail=f"الدين رقم {item.debt_id} غير موجود")
            if debt.customer_name != data.customer_name:
                raise HTTPException(status_code=400, detail=f"الدين رقم {item.debt_id} لا يخص هذا الزبون")
            if debt.status == "paid":
                raise HTTPException(status_code=400, detail=f"القائمة {debt.invoice_number} مسددة بالكامل مسبقاً")
            if item.amount_applied <= 0:
                raise HTTPException(status_code=400, detail="المبلغ المطبق يجب أن يكون أكبر من صفر")
            if item.amount_applied > debt.remaining_amount:
                raise HTTPException(
                    status_code=400,
                    detail=f"المبلغ المطبق ({item.amount_applied}) أكبر من المبلغ المتبقي ({debt.remaining_amount}) للقائمة {debt.invoice_number}",
                )
            total_selected += item.amount_applied
            debt_objects.append((debt, item.amount_applied))
    else:
        # No invoice selected: a manual/unallocated payment on account.
        if not data.manual_amount or data.manual_amount <= 0:
            raise HTTPException(
                status_code=400,
                detail="يرجى اختيار قائمة واحدة على الأقل أو إدخال مبلغ يدوي أكبر من صفر",
            )
        total_selected = data.manual_amount

    # Calculate net received
    net_received = total_selected - data.discount - data.returns_amount
    if net_received < 0:
        raise HTTPException(status_code=400, detail="المبلغ الصافي لا يمكن أن يكون سالباً")

    # A rep collecting cash out in the field holds it until accounting
    # confirms receipt (see /payments/confirm-handover); anyone else
    # recording a payment is assumed to already be at the accounting side.
    role = await resolve_user_role(db, current_user)
    is_rep = role == "rep"
    now = datetime.now(timezone.utc)

    # Create payment record
    payment = Payments(
        user_id=str(current_user.id),
        customer_name=data.customer_name,
        total_selected=total_selected,
        discount=data.discount,
        returns_amount=data.returns_amount,
        net_received=net_received,
        notes=data.notes,
        rep_name=current_user.name or current_user.email,
        status="active",
        handed_over=not is_rep,
        handed_over_at=None if is_rep else now,
        handed_over_by=None if is_rep else str(current_user.id),
    )
    db.add(payment)
    await db.flush()  # Get payment.id

    # Create payment details and update debts
    for debt, amount_applied in debt_objects:
        # Create detail record
        detail = PaymentDetails(
            user_id=str(current_user.id),
            payment_id=payment.id,
            debt_id=debt.id,
            amount_applied=amount_applied,
        )
        db.add(detail)

        # Update debt
        new_remaining = debt.remaining_amount - amount_applied
        if new_remaining <= 0.01:  # Float tolerance
            debt.remaining_amount = 0
            debt.status = "paid"
        else:
            debt.remaining_amount = round(new_remaining, 2)
            debt.status = "partial"

    await db.commit()
    await db.refresh(payment)

    logger.info(f"Payment {payment.id} processed: {total_selected} - {data.discount} - {data.returns_amount} = {net_received}")
    return payment


@router.post("/payments/{payment_id}/cancel", response_model=PaymentResponse)
async def cancel_payment(
    payment_id: int,
    data: CancelPaymentRequest,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Cancel a previously recorded payment, restoring any debts it was applied to.
    Requires can_delete permission on the debts page (matches the frontend's
    canDelete('debts') gate on this action)."""
    from sqlalchemy import select

    await require_permission(db, current_user, "debts", "delete")

    result = await db.execute(select(Payments).where(Payments.id == payment_id))
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")
    if payment.status == "canceled":
        raise HTTPException(status_code=400, detail="هذه الدفعة ملغاة مسبقاً")

    details_result = await db.execute(
        select(PaymentDetails).where(PaymentDetails.payment_id == payment_id)
    )
    details = details_result.scalars().all()

    service = DebtsService(db)
    for detail in details:
        debt = await service.get_by_id(detail.debt_id)
        if not debt:
            # Debt was deleted separately; nothing to restore for this portion.
            logger.warning(f"Cancel payment {payment_id}: debt {detail.debt_id} no longer exists")
            continue
        new_remaining = min(debt.remaining_amount + detail.amount_applied, debt.amount)
        new_remaining = max(new_remaining, 0)
        debt.remaining_amount = round(new_remaining, 2)
        if debt.remaining_amount >= debt.amount - 0.01:
            debt.status = "unpaid"
        elif debt.remaining_amount <= 0.01:
            debt.status = "paid"
        else:
            debt.status = "partial"

    payment.status = "canceled"
    payment.canceled_at = datetime.now(timezone.utc)
    payment.canceled_by = current_user.name or current_user.email
    payment.cancel_reason = data.reason

    await db.commit()
    await db.refresh(payment)

    logger.info(f"Payment {payment_id} canceled by {payment.canceled_by}")
    return payment


@router.get("/payments", response_model=PaymentListResponse)
async def list_payments(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=2000),
    customer_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """List payment transactions"""
    from sqlalchemy import select, func

    query = select(Payments)
    count_query = select(func.count(Payments.id))

    if customer_name:
        query = query.where(Payments.customer_name == customer_name)
        count_query = count_query.where(Payments.customer_name == customer_name)

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    result = await db.execute(query.order_by(Payments.id.desc()).offset(skip).limit(limit))
    items = result.scalars().all()

    return {"items": items, "total": total, "skip": skip, "limit": limit}


# ---------- Cash Handover (rep -> accounting) ----------
@router.get("/payments/handover-summary", response_model=HandoverSummaryResponse)
async def get_handover_summary(
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Breaks settled payments down into 'still with the rep' (collected but
    not yet handed to accounting) vs 'confirmed received by accounting',
    grouped by rep. Cancelled payments are excluded entirely."""
    from sqlalchemy import select

    result = await db.execute(select(Payments).order_by(Payments.created_at.desc()))
    all_payments = result.scalars().all()

    pending_total = 0.0
    confirmed_total = 0.0
    reps: dict[str, HandoverRepSummary] = {}

    for p in all_payments:
        if p.status == "canceled":
            continue
        if p.handed_over:
            confirmed_total += p.net_received
            continue
        pending_total += p.net_received
        key = p.user_id
        if key not in reps:
            reps[key] = HandoverRepSummary(
                user_id=p.user_id,
                rep_name=p.rep_name or "غير معروف",
                pending_total=0.0,
                payments=[],
            )
        reps[key].pending_total += p.net_received
        reps[key].payments.append(p)

    return HandoverSummaryResponse(
        reps=sorted(reps.values(), key=lambda r: -r.pending_total),
        pending_total=pending_total,
        confirmed_total=confirmed_total,
    )


@router.post("/payments/confirm-handover")
async def confirm_handover(
    request: ConfirmHandoverRequest,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Accounting confirms it physically received cash a rep collected —
    marks the given payments as handed_over, zeroing them out of that rep's
    pending total. Supports confirming one payment or many (select-all) in
    a single call. Requires can_delete on the debts page, same gate as
    cancel_payment."""
    await require_permission(db, current_user, "debts", "delete")

    if not request.payment_ids:
        raise HTTPException(status_code=400, detail="يرجى تحديد دفعة واحدة على الأقل")

    from sqlalchemy import select

    result = await db.execute(select(Payments).where(Payments.id.in_(request.payment_ids)))
    payments = result.scalars().all()

    now = datetime.now(timezone.utc)
    confirmed_count = 0
    for p in payments:
        if p.status == "canceled" or p.handed_over:
            continue
        p.handed_over = True
        p.handed_over_at = now
        p.handed_over_by = str(current_user.id)
        confirmed_count += 1

    await db.commit()
    logger.info(f"Confirmed handover for {confirmed_count} payments by {current_user.id}")
    return {"success": True, "confirmed_count": confirmed_count}


# ---------- Payment Receipt Photo (archive) ----------
@router.post("/payments/{payment_id}/receipt-image")
async def upload_receipt_image(
    payment_id: int,
    data: ReceiptImageUpload,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Attach a captured (pharmacy-stamped) receipt photo to a payment record."""
    from sqlalchemy import select

    if not data.image_base64:
        raise HTTPException(status_code=400, detail="الصورة مطلوبة")
    if len(data.image_base64) > MAX_RECEIPT_IMAGE_CHARS:
        raise HTTPException(status_code=400, detail="حجم الصورة كبير جداً")

    result = await db.execute(select(Payments).where(Payments.id == payment_id))
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")

    payment.receipt_image = data.image_base64
    payment.receipt_captured_at = datetime.now(timezone.utc)
    await db.commit()

    logger.info(f"Receipt image saved for payment {payment_id}")
    return {"success": True}


@router.get("/payments/{payment_id}/receipt-image")
async def get_receipt_image(
    payment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserResponse = Depends(get_current_user),
):
    """Fetch the archived receipt photo for a payment, if one was captured."""
    from sqlalchemy import select

    result = await db.execute(select(Payments).where(Payments.id == payment_id))
    payment = result.scalar_one_or_none()
    if not payment or not payment.receipt_image:
        raise HTTPException(status_code=404, detail="لا توجد صورة وصل لهذه الدفعة")

    return {
        "payment_id": payment.id,
        "image_base64": payment.receipt_image,
        "captured_at": payment.receipt_captured_at,
    }